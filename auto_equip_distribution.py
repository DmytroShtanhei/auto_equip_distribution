"""
Script
for processing spreadsheet "Договір" (of file "Рознарядка.xlsx") and file "Групування.xlsx"
and adding new spreadsheets "Групування" and "Рознарядка" (to file "Рознарядка.xlsx")
"""
import copy
from openpyxl import load_workbook
import datetime
import lvu_names
import utils
from named_styles import header_style, data_style
import locale

locale.setlocale(locale.LC_ALL, "")

distribution_wb = load_workbook(filename='Рознарядка.xlsx')
contract_ws = distribution_wb['Договір']

original_grouping_wb = load_workbook(filename='Групування.xlsx')
original_grouping_ws = original_grouping_wb.active

if 'Групування' in distribution_wb:
    del distribution_wb['Групування']

grouping_copied_ws = distribution_wb.create_sheet('Групування')
utils.copy_table(source_ws=original_grouping_ws, target_ws=grouping_copied_ws)
utils.customize_grouping_copied_ws(grouping_copied_ws=grouping_copied_ws)
utils.prepare_grouping_table(grouping_ws_to_be_prepared=grouping_copied_ws)
utils.style_table_in_worksheet(workbook=distribution_wb,
                               worksheet=grouping_copied_ws,
                               custom_header_style=header_style,
                               custom_data_style=data_style,
                               max_header_row=4,
                               )

if 'Рознарядка. Перевірка' in distribution_wb:
    del distribution_wb['Рознарядка. Перевірка']

distribution_ws = distribution_wb.create_sheet('Рознарядка. Перевірка')

if 'Рознарядка по регіонах' in distribution_wb:
    del distribution_wb['Рознарядка по регіонах']

# Validate data cells in original worksheets (contract_ws and original_grouping_ws).
contract_ws_is_valid = utils.is_contract_ws_valid(contract_ws)
original_grouping_ws_is_valid = utils.is_original_grouping_ws_valid(original_grouping_ws)
# Generate error message in distribution_ws and exit script if data in original worksheets aren't valid
utils.validation_error_message_to_distribution_ws(distribution_wb,
                                                  distribution_ws,
                                                  contract_ws_is_valid,
                                                  original_grouping_ws_is_valid)

# Get list of positions from Contract table
positions_n_units_list = utils.get_positions_n_units_list(contract_ws)
# Get list of LVU from Grouping table
lvu_list = utils.get_lvu_list(grouping_copied_ws)

# Get distribution data list
distribution_data_list = utils.get_distribution_data_list(positions_n_units_list,
                                                          grouping_copied_ws,
                                                          )

# get_distribution_full_list
distribution_full_list = utils.get_distribution_full_list(positions_n_units_list,
                                                          lvu_list,
                                                          distribution_data_list,
                                                          )

# Replace LVU codes in distribution_full_list with LVU names
utils.replace_lvu_codes_with_names(distribution_full_list, lvu_names.lvu_names_list)

# distribution_full_list_sorted_by_lvu = sorted(distribution_full_list, key=itemgetter(0))

# Sort rows by LVU
# great explanation is here: https://stackoverflow.com/questions/36770509/sorting-with-two-key-arguments
# (strxfrm() is used for locale aware sorting)
distribution_full_list_sorted_by_lvu = sorted(distribution_full_list, key=lambda element: (locale.strxfrm(element[0])))
# Insert Numbers by Order as first item for each row
distribution_full_list_sorted_by_lvu_with_nbo = copy.deepcopy(distribution_full_list_sorted_by_lvu)
counter = 1
for item in distribution_full_list_sorted_by_lvu_with_nbo:
    item.insert(0, counter)
    counter += 1
# print(*distribution_full_list_sorted_by_lvu, sep='\n')

# Create header for distribution spreadsheet
utils.create_header_for_distribution_ws(positions_n_units_list,
                                        distribution_ws)

# Populate distribution_ws with distribution data from distribution_full_list
utils.append_list_to_worksheet(distribution_full_list_sorted_by_lvu_with_nbo,
                               distribution_ws)

# Style the table in distribution spreadsheet
utils.style_table_in_worksheet(workbook=distribution_wb,
                               worksheet=distribution_ws,
                               custom_header_style=header_style,
                               custom_data_style=data_style,
                               max_header_row=2,
                               )

# print(utils.get_units_for_position(grouping_copied_ws, 8))
# Add check sums to distribution spreadsheet
utils.add_distribution_check_sum(distribution_ws,
                                 grouping_copied_ws,
                                 lvu_list,
                                 positions_n_units_list,
                                 distribution_wb,
                                 data_style)

# Highlight problems with Distribution check sums and get status of correctness of sums
sums_are_correct = utils.check_n_highlight_distribution_sums(contract_ws,
                                                             distribution_ws,
                                                             lvu_list,
                                                             positions_n_units_list)

# Highlight problems with Grouping check sums
utils.check_n_highlight_grouping_sums(distribution_ws,
                                      contract_ws,
                                      grouping_copied_ws,
                                      lvu_list,
                                      positions_n_units_list)

# Highlight problems with units
utils.check_n_highlight_grouping_units(distribution_ws,
                                       lvu_list,
                                       positions_n_units_list)

# ---------------- Create Datasheet with Distribution list grouped by Regions ----------------

# Extend distribution full list with information about the region for each LVU
distribution_full_list_extended = utils.get_extend_distribution_full_list(distribution_full_list,
                                                                          lvu_names.lvu_names_list)

# Form list of Distribution Lists grouped by regions
grouped_by_region_list_with_nbo = utils.form_grouped_by_region_list(distribution_full_list_extended)

# Create new list 'Рознарядка по регіонах' in distribution_wb

distribution_by_region_ws = distribution_wb.create_sheet('Рознарядка по регіонах')

# Create header for Distribution by Regions spreadsheet
utils.create_header_for_distribution_ws(positions_n_units_list,
                                        distribution_by_region_ws)

# Populate distribution_ws with distribution data from distribution_full_list
utils.append_list_to_worksheet(grouped_by_region_list_with_nbo,
                               distribution_by_region_ws)

# Style table in Distribution by Region Worksheet
utils.style_table_in_worksheet(workbook=distribution_wb,
                               worksheet=distribution_by_region_ws,
                               custom_header_style=header_style,
                               custom_data_style=data_style,
                               max_header_row=2)

# Customize look of Distribution by Region Table
utils.customize_grouped_by_region_table(distribution_by_region_ws, lvu_names.lvu_names_list, sums_are_correct)

# Make given sheet active
distribution_wb.active = distribution_wb['Рознарядка. Перевірка']
# Save distribution workbook
distribution_wb.save(f'Рознарядка {datetime.datetime.now().strftime("%Y-%m-%d_T%H%M%S")}.xlsx')
