"""
Utility functions
for processing "Договір" and "Групування" .xlsx source files
and creating "Рознарядка" .xlsx file
"""
import copy
from decimal import Decimal
from openpyxl.styles import PatternFill, Font
import locale


def copy_table(source_ws, target_ws):
    """Copy table from Source Spreadsheet to Target Spreadsheet of Another Workbook."""
    # Get source table as a list
    source_table_as_list = []
    for row in source_ws.rows:
        row_as_list = []
        for cell in row:
            row_as_list.append(cell.value)
        source_table_as_list.append(row_as_list)

    # Paste table as a list to target worksheet
    append_list_to_worksheet(source_table_as_list, target_ws)


def customize_grouping_copied_ws(grouping_copied_ws):
    """Customize Copied to Distribution Workbook Grouping Spreadsheet so it would look like the or original one """
    grouping_copied_ws.merge_cells('A1:A4')
    grouping_copied_ws.merge_cells('B1:B4')
    grouping_copied_ws.merge_cells('C1:F2')
    grouping_copied_ws.merge_cells('C3:C4')
    grouping_copied_ws.merge_cells('D3:D4')
    grouping_copied_ws.merge_cells('E3:F4')
    grouping_copied_ws.merge_cells('G1:G4')
    grouping_copied_ws.merge_cells('H1:H2')
    grouping_copied_ws.merge_cells('H3:H4')
    grouping_copied_ws.merge_cells('I1:I4')
    grouping_copied_ws.merge_cells('J1:J4')
    grouping_copied_ws.merge_cells('K1:K4')
    grouping_copied_ws.merge_cells('L1:L4')
    grouping_copied_ws.merge_cells('M1:M4')


def prepare_grouping_table(grouping_ws_to_be_prepared):
    """Code for automated preparing of copied grouping spreadsheet."""
    # Spread names of position inside each group of formerly merged cells in the position colon.
    for col in grouping_ws_to_be_prepared.iter_cols(min_col=13, max_col=13, min_row=5):
        previous_cell_val = 'Не визначено'
        for cell in col:
            if cell.value is None:
                cell.value = previous_cell_val
            previous_cell_val = cell.value


def get_positions_n_units_list(contract_ws):
    """Get position list from contract spreadsheet."""
    contract_positions_n_units_list = []
    for row in contract_ws.rows:
        index_row = row[0].row
        if index_row >= 3:
            position_n_unit = [row[0].value, row[14].value]
            contract_positions_n_units_list.append(position_n_unit)

    return contract_positions_n_units_list


def get_lvu_list(grouping_ws):
    """Get LVU list from distribution spreadsheet."""
    distribution_lvu_list = []
    for col in grouping_ws.iter_cols(min_row=5, min_col=8, max_col=8, values_only=True):
        for val in col:
            if val is not None:
                distribution_lvu_list.append(val)

    return sorted(list(set(distribution_lvu_list)))


# def check_grouping_positions(contract_ws, original_grouping_ws):
#     """Compare positions in Original Grouping Workbook with positions in Contract Worksheet"""
#     pass
#     # Get Contract positions set
#     for col in contract_ws.iter_cols(min_row=3, max_col=1, values_only=True):
#         contract_pos_set = set()
#         for cell in col:
#             contract_pos_set.add(cell)
#         print(contract_pos_set)
#
#     # Get Original Grouping position set
#     for col in original_grouping_ws.iter_cols(min_row=5, min_col=13, max_col=13, values_only=False):
#         grouping_pos_set = set()
#         for cell in col:
#             if not isinstance(cell, MergedCell):
#                 grouping_pos_set.add(cell.value)
#         print(grouping_pos_set)


def get_lvu_list_for_position(grouping_ws, position):
    """Get list of LVU which ordered given position."""
    lvu_list_for_pos = []
    for row in grouping_ws.iter_rows(min_row=5, min_col=2, max_col=13, values_only=True):
        if row[11] == position:
            lvu_list_for_pos.append(row[6])

    return sorted(list(set(lvu_list_for_pos)))


def get_distribution_data_list(positions_n_units_list, grouping_ws):
    """Create list with raw data of sum per position per lvu: lvu | position | sum"""
    distribution_data_list = []
    for item in positions_n_units_list:
        lvu_list_for_position = get_lvu_list_for_position(grouping_ws=grouping_ws, position=item[0])
        for curr_lvu in lvu_list_for_position:
            dist_data_row = []
            curr_sum = 0
            for row in grouping_ws.iter_rows(min_row=5, min_col=2, max_col=13, values_only=True):
                if row[11] == item[0] and row[6] == curr_lvu:
                    curr_sum += Decimal(str(row[2]))
            # print(f'{curr_lvu} {item[0]} {curr_sum}')
            dist_data_row.append(curr_lvu)
            dist_data_row.append(item[0])
            dist_data_row.append(curr_sum)
            # print(dist_data_row)
            distribution_data_list.append(dist_data_row)

    return distribution_data_list


def get_sum_from_distribution_data_list(distribution_data_list, lvu, position):
    """Get needed sum from raw data list for given LVU and position."""
    # print(distribution_data_list)
    for item in distribution_data_list:
        # print(item)
        if lvu == item[0] and position == item[1]:
            return item[2]

    return 0


def get_distribution_full_list(positions_n_units_list, lvu_list, distribution_data_list):
    """
    Create distribution data full list in form of asking table:
    lvu | sum_for_pos_1 | ... | sum_for_pos_n
    """
    distribution_full_list = []
    for curr_lvu in lvu_list:
        curr_lvu_list = [curr_lvu]
        for item in positions_n_units_list:
            sum_for_lvu_for_pos = get_sum_from_distribution_data_list(distribution_data_list, curr_lvu, item[0])
            curr_lvu_list.append(sum_for_lvu_for_pos)
        distribution_full_list.append(curr_lvu_list)

    return distribution_full_list


def replace_lvu_codes_with_names(distribution_full_list, lvu_names_list):
    """Replace codes in distribution_full_list with names from lvu_names_list."""
    for curr_lvu_list in distribution_full_list:
        curr_lvu_code = str(curr_lvu_list[0])
        for item in lvu_names_list:
            i = lvu_names_list.index(item)
            if item[0] == curr_lvu_code:
                curr_lvu_list[0] = lvu_names_list[i][2]


def create_header_for_distribution_ws(positions_n_units_list, distribution_ws):
    """
    Create header for distribution spreadsheet in form of asked table:
    lvu | sum_for_pos_1 | ... | sum_for_pos_n
    """
    # Initialise header list
    distribution_header_list = [
        '№ п-п',
        'Назва ЛВУМГ (ЛВУМГ що замовляло)',
    ]
    # Append positions to header list
    for item in positions_n_units_list:
        distribution_header_list.append(item[0])

    # Populate distribution_ws header row
    for row in distribution_ws.iter_rows(max_row=1, max_col=len(distribution_header_list)):
        p = 0
        for cell in row:
            cell.value = distribution_header_list[p]
            p += 1

    # Populate units row
    for row in distribution_ws.iter_rows(min_row=2, max_row=2, min_col=3, max_col=len(positions_n_units_list) + 2):
        p = 0
        for cell in row:
            cell.value = positions_n_units_list[p][1]
            p += 1

    # Merge appropriate cells
    distribution_ws.merge_cells('A1:A2')
    distribution_ws.merge_cells('B1:B2')


# def init_table_in_distribution_ws(positions_n_units_list, lvu_list, distribution_ws):
#     """
#     Create header for distribution spreadsheet in form of asked table:
#     lvu | sum_for_pos_1 | ... | sum_for_pos_n
#     """
#     # Initialise header list
#     distribution_header_list = [
#         '№ п-п',
#         'Назва ЛВУМГ (ЛВУМГ що замовляло)',
#     ]
#     # Append positions to header list
#     for item in positions_n_units_list:
#         distribution_header_list.append(item[0])
#
#     # Populate distribution_ws header row
#     for row in distribution_ws.iter_rows(max_row=1, max_col=len(distribution_header_list)):
#         p = 0
#         for cell in row:
#             cell.value = distribution_header_list[p]
#             p += 1
#
#     # Populate units row
#     for row in distribution_ws.iter_rows(min_row=2, max_row=2, min_col=3, max_col=len(positions_n_units_list) + 2):
#         p = 0
#         for cell in row:
#             cell.value = positions_n_units_list[p][1]
#             p += 1
#
#     # Populate Number in order column
#     for col in distribution_ws.iter_cols(max_col=1, min_row=3, max_row=len(lvu_list) + 2):
#         num_in_order = 1
#         for cell in col:
#             cell.value = num_in_order
#             num_in_order += 1
#
#     # Merge appropriate cells
#     distribution_ws.merge_cells('A1:A2')
#     distribution_ws.merge_cells('B1:B2')


def append_list_to_worksheet(list_of_data,
                             worksheet):
    """Add data from List of Data to Worksheet."""
    for row in list_of_data:
        worksheet.append(row)


# def populate_table_in_distribution_ws(distribution_ws, lvu_list, positions_n_units_list, distribution_full_list):
#     """Populate distribution_ws with distribution data from distribution_full_list."""
#     curr_lvu_list_index = 0
#     for row in distribution_ws.iter_rows(min_row=3,
#                                          max_row=len(lvu_list) + 2,
#                                          max_col=len(positions_n_units_list) + 2,
#                                          ):
#         curr_item_in_lvu_list_index = 0
#         for cell in row:
#             value = distribution_full_list[curr_lvu_list_index][curr_item_in_lvu_list_index]
#             if value == 0:
#                 cell.value = None
#             else:
#                 cell.value = value
#             curr_item_in_lvu_list_index += 1
#         curr_lvu_list_index += 1


# def auto_adjust_col_width(worksheet):
#     """Change columns width according to columns value length"""
#     for col in worksheet.columns:
#         max_length = 0
#         column = col[0].column_letter  # Get the column name
#         for cell in col:
#             try:  # Necessary to avoid error on empty cells
#                 if len(str(cell.value)) > max_length:
#                     max_length = len(str(cell.value))
#             except:
#                 pass
#         adjusted_width = (max_length + 2)
#         worksheet.column_dimensions[column].width = adjusted_width


def style_table_in_worksheet(workbook, worksheet, custom_header_style, custom_data_style, max_header_row=1):
    """Style the table in a given worksheet."""
    # Register custom Named Styles in the Workbook if they are not registered yet.
    if custom_header_style.name not in workbook.named_styles:
        workbook.add_named_style(custom_header_style)
    if custom_data_style.name not in workbook.named_styles:
        workbook.add_named_style(custom_data_style)

    # Add Named Styles to cells
    for row in worksheet.rows:
        row_index = row[0].row
        if row_index <= max_header_row:
            for cell in row:
                cell.style = custom_header_style.name
        else:
            for cell in row:
                cell.style = custom_data_style.name
                # if cell.column > 2:
                #     cell.number_format = '0.000'

    # Change column width
    worksheet.column_dimensions['B'].width = 37 + 2


def get_quantity_sum_formula_for_position(grouping_ws, position):
    """
    Get string that represents range of cells (in form like this: 'D5:D100')
    with quantity needed for given position.
    """
    row_range_index_list = []
    for col in grouping_ws.iter_cols(min_row=5, min_col=13, max_col=13):
        for cell in col:
            if cell.value == position:
                row_range_index_list.append(cell.row)
    if not row_range_index_list:
        return 'поз. відс.'
    else:
        return f'=SUM(Групування!D{min(row_range_index_list)}:D{max(row_range_index_list)})'


def get_units_for_position(grouping_ws, position):
    """Get string that represents list of units found for given position."""
    units_list = []
    for col in grouping_ws.iter_cols(min_row=5, min_col=13, max_col=13):
        for cell in col:
            if cell.value == position:
                units_list.append(grouping_ws[f'C{cell.row}'].value)
    unique_units_list = list(set(units_list))
    unique_units_str = ''
    for unit in unique_units_list:
        unique_units_str += f'{str(unit)}, '
    return unique_units_str[:-2]


def add_distribution_check_sum(distribution_ws, grouping_ws, lvu_list, positions_n_units_list, distribution_wb,
                               custom_data_style):
    """
    Add check sums and units for each position for distribution, grouping and contract
    to distribution spreadsheet
    """
    # Register custom Named Styles in the Workbook if they are not registered yet.
    if custom_data_style.name not in distribution_wb.named_styles:
        distribution_wb.add_named_style(custom_data_style)

    # Add check sum for distribution table
    check_sum_row_index = len(lvu_list) + 4
    for row in distribution_ws.iter_rows(min_row=check_sum_row_index, max_row=check_sum_row_index):
        row[1].value = 'Рознарядка. Сумарна кількість:'
        row[1].style = custom_data_style.name
        row[1].font = Font(bold=True)
        for i in range(2, len(positions_n_units_list) + 2):
            row[i].value = f'=SUM({row[i].column_letter}{3}:{row[i].column_letter}{len(lvu_list) + 2})'
            row[i].style = custom_data_style.name
            row[i].font = Font(bold=True)

    # Add check sum for contract
    check_sum_row_index = len(lvu_list) + 6
    for row in distribution_ws.iter_rows(min_row=check_sum_row_index, max_row=check_sum_row_index):
        row[1].value = 'Договір. Сумарна кількість:'
        row[1].style = custom_data_style.name
        row[1].font = Font(bold=True)
        for i in range(2, len(positions_n_units_list) + 2):
            row[i].value = f'=Договір!P{i + 1}'
            row[i].style = custom_data_style.name
            row[i].font = Font(bold=True)
    # Add units for contract check sum
    check_sum_row_index = len(lvu_list) + 7
    for row in distribution_ws.iter_rows(min_row=check_sum_row_index, max_row=check_sum_row_index):
        row[1].value = 'Договір. Одиниці виміру:'
        row[1].style = custom_data_style.name
        row[1].font = Font(bold=True)
        for i in range(2, len(positions_n_units_list) + 2):
            row[i].value = f'=Договір!O{i + 1}'
            row[i].style = custom_data_style.name
            row[i].font = Font(bold=True)

    # Add check sum for grouping
    check_sum_row_index = len(lvu_list) + 9
    for row in distribution_ws.iter_rows(min_row=check_sum_row_index, max_row=check_sum_row_index):
        row[1].value = 'Групування. Сумарна кількість:'
        row[1].style = custom_data_style.name
        for i in range(2, len(positions_n_units_list) + 2):
            row[i].value = get_quantity_sum_formula_for_position(grouping_ws, positions_n_units_list[i - 2][0])
            row[i].style = custom_data_style.name
    # Add units for grouping check sum
    check_sum_row_index = len(lvu_list) + 10
    for row in distribution_ws.iter_rows(min_row=check_sum_row_index, max_row=check_sum_row_index):
        row[1].value = 'Групування. Одиниці виміру:'
        row[1].style = custom_data_style.name
        for i in range(2, len(positions_n_units_list) + 2):
            row[i].value = get_units_for_position(grouping_ws, positions_n_units_list[i - 2][0])
            row[i].style = custom_data_style.name


def check_n_highlight_distribution_sums(contract_ws,
                                        distribution_ws,
                                        lvu_list,
                                        positions_n_units_list):
    """
    Highlight cells with Distribution Check Sums that don't correspond to Contract Sums.
    Return status of correctness of sums.
    """
    sums_are_correct = True

    curr_position = 1
    for col in distribution_ws.iter_cols(min_col=3, max_col=len(positions_n_units_list) + 2):
        # Get Distribution Check Sum for Position from Distribution Table
        distribution_sum = 0
        for i in range(2, len(lvu_list) + 2):
            val = col[i].value

            if val is None:
                val = 0
            # print(Decimal(str(val)))
            # print(val)
            distribution_sum += val

        # Get Contract Sum for Position from Contract Table
        row_index = curr_position + 2
        contract_sum = Decimal(str(contract_ws[f'P{row_index}'].value))
        curr_position += 1

        # Highlight distribution sums that don't correspond to contract units
        if distribution_sum != contract_sum:
            col[len(lvu_list) + 3].fill = PatternFill(fill_type='solid', start_color='00FF0000')

            distribution_ws.cell(len(lvu_list) + 4, len(positions_n_units_list) + 3).value = \
                ' <- Сумарна кількість не відповідає Договору (можливі причини див. нижче)'
            distribution_ws.cell(len(lvu_list) + 4, len(positions_n_units_list) + 3).font = Font(color='00FF0000')
            sums_are_correct = False

    return sums_are_correct


def check_n_highlight_grouping_sums(distribution_ws,
                                    contract_ws,
                                    grouping_copied_ws,
                                    lvu_list,
                                    positions_n_units_list):
    """Highlight cells with Grouping sums that don't correspond to Contract sums"""
    for col in distribution_ws.iter_cols(min_col=3, max_col=len(positions_n_units_list) + 2):
        # Get Grouping Check Sum for Position from Grouping Table (Grouping worksheet)
        curr_position = col[0].value
        grouping_sum = 0
        for row in grouping_copied_ws.rows:
            if row[12].value == curr_position:
                grouping_sum += Decimal(str(row[3].value))

        # Get Contract Sum for Position from Contract Table
        contract_sum = 0
        for row in contract_ws.iter_rows(min_row=3):
            if row[0].value == curr_position:
                contract_sum = Decimal(str(row[15].value))

        # Highlight Grouping Sums that don't correspond to Contract Sums
        if grouping_sum != contract_sum:
            col[len(lvu_list) + 8].fill = PatternFill(fill_type='solid', start_color='00FF9900')
            distribution_ws.cell(len(lvu_list) + 9, len(positions_n_units_list) + 3).value = \
                ' <- Сумарна кількість не відповідає Договору. Відкоригуйте ФАЙЛ "Групування.xlsx"'
            distribution_ws.cell(len(lvu_list) + 9, len(positions_n_units_list) + 3).font = Font(color='00FF9900')


def check_n_highlight_grouping_units(distribution_ws,
                                     lvu_list,
                                     positions_n_units_list):
    """Highlight cells with Grouping units that don't correspond to Contract units"""
    for col in distribution_ws.iter_cols(min_col=3, max_col=len(positions_n_units_list) + 2):
        grouping_units = col[len(lvu_list) + 9].value
        distribution_units = col[1].value
        # Highlight grouping units that don't correspond to contract units
        if grouping_units.strip('.').lower() != distribution_units.strip('.').lower():
            col[len(lvu_list) + 9].fill = PatternFill(fill_type='solid', start_color='00FF9900')
            distribution_ws.cell(len(lvu_list) + 10, len(positions_n_units_list) + 3).value = \
                ' <- Одиниці виміру не відповідають Договору'
            distribution_ws.cell(len(lvu_list) + 10, len(positions_n_units_list) + 3).font = Font(color='00FF9900')


def get_extend_distribution_full_list(distribution_full_list, lvu_names_list):
    """Extend distribution full list with information about the region for each LVU"""
    distribution_full_list_extended = copy.deepcopy(distribution_full_list)
    for lvu_sums in distribution_full_list_extended:
        lvu_sums_name = lvu_sums[0]
        for lvu_names in lvu_names_list:
            lvu_names_name = lvu_names[2]
            if lvu_sums_name == lvu_names_name:
                lvu_sums.append(lvu_names[3])

    return distribution_full_list_extended


def get_distribution_list_for_region(distribution_full_list_extended, region_name):
    """
    Form Distribution list for given region,
    add numbers by order,
    add row with totals per position,
    add empty row for nicer look :)
    """
    # Form Distribution list for given region
    distribution_one_region_list = []
    for lvu_sums in distribution_full_list_extended:
        if lvu_sums[-1] == region_name:
            distribution_one_region_list.append(lvu_sums)
    distribution_one_region_list_sorted = sorted(distribution_one_region_list,
                                                 key=lambda element: (locale.strxfrm(element[0])))

    # Prepend numbers by order
    counter = 1
    for lvu_sums in distribution_one_region_list_sorted:
        lvu_sums.insert(0, counter)
        counter += 1

    # Form row with total numbers per position with name of region as second item
    total_per_position_list = ['', region_name]
    for i in range(len(distribution_one_region_list_sorted[0]) - 3):

        # Get total for given position (i + 2)
        total = 0
        for item in distribution_one_region_list_sorted:
            total += item[i + 2]

        total_per_position_list.append(total)

    # Form empty lvu_sums
    empty_lvu_sums = []
    for i in range(len(distribution_one_region_list_sorted[0])):
        empty_lvu_sums.append('')

    # Append Distribution list for given region with empty lvu_sums
    distribution_one_region_list_sorted.append(total_per_position_list)
    distribution_one_region_list_sorted.append(empty_lvu_sums)

    return distribution_one_region_list_sorted


def form_grouped_by_region_list(distribution_full_list_extended):
    """Form list of Distribution Lists grouped by regions"""
    # Get set of regions
    grouped_by_region_list = []
    regions_list = []
    for lvu_sums in distribution_full_list_extended:
        regions_list.append(lvu_sums[-1])
    regions_list_sorted = sorted(list(set(regions_list)))

    # Group LVU list by region
    for region in regions_list_sorted:
        the_region_distribution_list = get_distribution_list_for_region(distribution_full_list_extended, region)
        grouped_by_region_list.extend(the_region_distribution_list)

    del grouped_by_region_list[-1]

    # Get rid of last element with information about the region for each LVU
    extended_length = len(grouped_by_region_list[0])
    for lvu_sums in grouped_by_region_list:
        if len(lvu_sums) == extended_length:
            del lvu_sums[-1]
    # print(*grouped_by_region_list, sep='\n')

    return grouped_by_region_list


def customize_grouped_by_region_table(distribution_by_region_ws, lvu_names_list, sums_are_correct):
    """Customize look of Distribution by Region Table"""
    # Customize rows
    for row in distribution_by_region_ws.iter_rows(min_row=3):

        # Choose and customize rows with totals
        if row[1].value in [item[-1] for item in lvu_names_list]:
            # Merge first and second cells and paste back value of the second cell
            temp = row[1].value
            distribution_by_region_ws.merge_cells(f'{row[0].coordinate}:{row[1].coordinate}')
            row[0].value = temp
            # Style cells
            for cell in row:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(fill_type='solid', start_color='E9E9E9')

        # Choose and merge cells in empty rows
        if row[1].value == '':
            distribution_by_region_ws.merge_cells(f'{row[0].coordinate}:{row[-1].coordinate}')

        # Highlight all rows if check sums in distribution list aren't correct
        if not sums_are_correct:
            for cell in row:
                cell.fill = PatternFill(fill_type='solid', start_color='00FF0000')

    if not sums_are_correct:
        distribution_by_region_ws.\
            append(['Сумарна кількість не відповідає Договору (можливі причини див. аркуш "Рознарядка. Перевірка")'])
        distribution_by_region_ws[f'A{distribution_by_region_ws.max_row}'].font = Font(color='00FF0000')
